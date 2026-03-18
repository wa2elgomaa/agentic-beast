"""Excel file processor for analytics report parsing and validation."""

import csv
from datetime import datetime
from io import BytesIO
from typing import List, Tuple

from openpyxl import load_workbook
from pydantic import ValidationError

from app.logging import get_logger
from app.schemas.ingestion import ExcelRow, RowError

logger = get_logger(__name__)

# Legacy fixed-schema mappings used by the old generic ingestion endpoints.
# Task-based ingestion now uses parse_tabular_rows(...) plus the task's schema mapping.
LEGACY_ANALYTICS_COLUMNS = {
    "date": "report_date",
    "platform": "platform",
    "profile_id": "profile_id",
    "profile_name": "profile_name",
    "reach": "reach",
    "impressions": "impressions",
    "engagement_rate": "engagement_rate",
    "likes": "likes",
    "comments": "comments",
    "shares": "shares",
    "saves": "saves",
}


class ExcelProcessor:
    """Processor for Excel analytics reports."""

    @staticmethod
    def parse_tabular_rows(file_data: bytes, filename: str, sheet_name: str = "Sheet1") -> Tuple[List[dict], List[RowError]]:
        """Parse a spreadsheet into raw source-column rows for task-based ingestion."""
        if filename.lower().endswith((".xlsx", ".xls")):
            return ExcelProcessor._parse_excel_rows_raw(file_data, sheet_name)
        if filename.lower().endswith(".csv"):
            return ExcelProcessor._parse_csv_rows_raw(file_data, sheet_name)
        return [], [RowError(row_number=0, error=f"Unsupported file type: {filename}")]

    @staticmethod
    def parse_excel(file_data: bytes, sheet_name: str = "Sheet1") -> Tuple[List[dict], List[RowError]]:
        """Parse a spreadsheet into the legacy fixed analytics schema.

        Args:
            file_data: Raw Excel file bytes.
            sheet_name: Name of sheet to parse.

        Returns:
            Tuple of (validated rows, errors).
        """
        validated_rows = []
        errors = []

        try:
            workbook = load_workbook(BytesIO(file_data))

            if sheet_name not in workbook.sheetnames:
                fallback_sheet = workbook.sheetnames[0] if workbook.sheetnames else None
                logger.warning(
                    "Sheet not found; falling back to first worksheet",
                    requested_sheet=sheet_name,
                    fallback_sheet=fallback_sheet,
                    available=workbook.sheetnames,
                )
                if not fallback_sheet:
                    return [], [RowError(row_number=0, error="Workbook contains no worksheets")]
                sheet_name = fallback_sheet

            worksheet = workbook[sheet_name]

            # Get headers from first row
            headers = []
            for cell in worksheet[1]:
                if cell.value:
                    headers.append(str(cell.value).lower())

            logger.info("Excel headers found", count=len(headers), headers=headers)

            # Validate headers
            if not headers:
                return [], [RowError(row_number=1, error="No headers found in Excel file")]

            # Map columns
            column_map = ExcelProcessor._map_columns(headers)
            if not column_map:
                logger.warning("Could not map Excel columns to expected schema")
                return [], [RowError(row_number=1, error="Column mapping failed")]

            # Process data rows (skip header row)
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Extract data using column map
                    row_data = {}
                    for excel_col, db_col in column_map.items():
                        col_index = headers.index(excel_col) if excel_col in headers else None
                        if col_index is not None:
                            value = row[col_index]
                            if isinstance(value, datetime):
                                value = value.date()
                            row_data[db_col] = value

                    # Add metadata
                    row_data["sheet_name"] = sheet_name
                    row_data["row_number"] = row_idx

                    # Validate row
                    try:
                        excel_row = ExcelRow(**row_data)
                        validated_rows.append(excel_row.model_dump())
                    except ValidationError as e:
                        logger.warning("Row validation failed", row_number=row_idx, errors=e.errors())
                        errors.append(
                            RowError(
                                row_number=row_idx,
                                error=f"Validation failed: {str(e.errors())[:100]}",
                            )
                        )

                except Exception as e:
                    logger.error("Error processing row", row_number=row_idx, error=str(e))
                    errors.append(RowError(row_number=row_idx, error=str(e)))

            logger.info(
                "Excel processing complete",
                valid_rows=len(validated_rows),
                error_rows=len(errors),
            )

            return validated_rows, errors

        except Exception as e:
            logger.error("Failed to parse Excel file", error=str(e))
            return [], [RowError(row_number=0, error=f"Excel parsing failed: {str(e)}")]

    @staticmethod
    def _map_columns(headers: List[str]) -> dict:
        """Map Excel columns to database columns.

        Args:
            headers: Column headers from Excel.

        Returns:
            Dict mapping Excel columns to DB columns.
        """
        column_map = {}

        for excel_col, db_col in LEGACY_ANALYTICS_COLUMNS.items():
            for header in headers:
                if excel_col.lower() in header.lower():
                    column_map[header] = db_col
                    break

        if not column_map:
            logger.warning("No columns could be mapped to expected schema")
            return {}

        logger.info("Column mapping complete", mapped_count=len(column_map))
        return column_map

    @staticmethod
    def _parse_excel_rows_raw(file_data: bytes, sheet_name: str = "Sheet1") -> Tuple[List[dict], List[RowError]]:
        """Parse Excel rows preserving raw source column names."""
        try:
            workbook = load_workbook(BytesIO(file_data))

            if sheet_name not in workbook.sheetnames:
                fallback_sheet = workbook.sheetnames[0] if workbook.sheetnames else None
                logger.warning(
                    "Sheet not found; falling back to first worksheet",
                    requested_sheet=sheet_name,
                    fallback_sheet=fallback_sheet,
                    available=workbook.sheetnames,
                )
                if not fallback_sheet:
                    return [], [RowError(row_number=0, error="Workbook contains no worksheets")]
                sheet_name = fallback_sheet

            worksheet = workbook[sheet_name]
            headers = [str(cell.value).lower().strip() if cell.value is not None else "" for cell in worksheet[1]]
            if not any(headers):
                return [], [RowError(row_number=1, error="No headers found in Excel file")]

            rows = []
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                row_data = {"sheet_name": sheet_name, "row_number": row_idx}
                for col_index, header in enumerate(headers):
                    if not header:
                        continue
                    row_data[header] = row[col_index] if col_index < len(row) else None
                rows.append(row_data)

            logger.info("Raw Excel rows parsed", sheet_name=sheet_name, row_count=len(rows))
            return rows, []
        except Exception as e:
            logger.error("Failed to parse Excel rows", error=str(e))
            return [], [RowError(row_number=0, error=f"Excel parsing failed: {str(e)}")]

    @staticmethod
    def _parse_csv_rows_raw(file_data: bytes, sheet_name: str = "Sheet1") -> Tuple[List[dict], List[RowError]]:
        """Parse CSV rows preserving raw source column names."""
        try:
            text = file_data.decode("utf-8-sig")
            reader = csv.DictReader(text.splitlines())
            if not reader.fieldnames:
                return [], [RowError(row_number=1, error="No headers found in CSV file")]

            normalized_headers = [str(header).lower().strip() for header in reader.fieldnames]
            rows = []
            for row_idx, raw_row in enumerate(reader, start=2):
                row_data = {"sheet_name": sheet_name, "row_number": row_idx}
                for original_header, normalized_header in zip(reader.fieldnames, normalized_headers):
                    if not normalized_header:
                        continue
                    row_data[normalized_header] = raw_row.get(original_header)
                rows.append(row_data)

            logger.info("Raw CSV rows parsed", row_count=len(rows))
            return rows, []
        except Exception as e:
            logger.error("Failed to parse CSV rows", error=str(e))
            return [], [RowError(row_number=0, error=f"CSV parsing failed: {str(e)}")]
