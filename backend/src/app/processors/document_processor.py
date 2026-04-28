"""Document processor for company document ingestion (PDF, Excel, plain text).

Extracts text from uploaded files, splits into chunks using LangChain's
RecursiveCharacterTextSplitter, and preserves metadata (filename, page/chunk index).
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from io import BytesIO
from typing import List, Optional

from app.config import settings

_logger = logging.getLogger(__name__)


@dataclass
class DocumentChunk:
    """A single processed chunk from a document."""
    text: str
    filename: str
    chunk_index: int
    page_number: Optional[int] = None
    metadata: dict = field(default_factory=dict)


class DocumentProcessor:
    """Processes uploaded files into text chunks for embedding and storage.

    Supports:
    - PDF (via pypdf)
    - Excel/CSV (via openpyxl)
    - Plain text (.txt, .md)
    """

    def __init__(
        self,
        chunk_size: Optional[int] = None,
        chunk_overlap: Optional[int] = None,
    ) -> None:
        self.chunk_size = chunk_size or getattr(settings, "document_chunk_size", 1000)
        self.chunk_overlap = chunk_overlap or getattr(settings, "document_chunk_overlap", 200)

    def process(self, file_data: bytes, filename: str) -> List[DocumentChunk]:
        """Extract text from file bytes and split into chunks.

        Args:
            file_data: Raw file bytes.
            filename: Original filename (used to determine parser).

        Returns:
            List of DocumentChunk objects ready for embedding.
        """
        fname_lower = filename.lower()

        if fname_lower.endswith(".pdf"):
            pages = self._extract_pdf(file_data, filename)
        elif fname_lower.endswith((".xlsx", ".xls", ".csv")):
            pages = self._extract_excel(file_data, filename)
        elif fname_lower.endswith((".txt", ".md")):
            pages = self._extract_text(file_data, filename)
        else:
            _logger.warning("Unsupported file type for document processor: %s", filename)
            pages = []

        return self._split_into_chunks(pages, filename)

    # ------------------------------------------------------------------
    # Private extraction methods
    # ------------------------------------------------------------------

    def _extract_pdf(self, data: bytes, filename: str) -> list[tuple[str, int]]:
        """Extract (text, page_number) tuples from a PDF file."""
        try:
            from pypdf import PdfReader  # pypdf is the maintained fork of PyPDF2
        except ImportError:
            try:
                from PyPDF2 import PdfReader  # fallback to PyPDF2
            except ImportError:
                _logger.error("pypdf/PyPDF2 not installed — cannot extract PDF text")
                return []

        pages: list[tuple[str, int]] = []
        try:
            reader = PdfReader(BytesIO(data))
            for page_num, page in enumerate(reader.pages, start=1):
                text = page.extract_text() or ""
                text = text.strip()
                if text:
                    pages.append((text, page_num))
        except Exception as exc:
            _logger.error("PDF extraction failed for %s: %s", filename, exc)
        return pages

    def _extract_excel(self, data: bytes, filename: str) -> list[tuple[str, int]]:
        """Extract text from Excel/CSV as concatenated row strings."""
        from io import StringIO
        import csv as csv_mod

        pages: list[tuple[str, int]] = []

        fname_lower = filename.lower()
        if fname_lower.endswith(".csv"):
            try:
                text_data = data.decode("utf-8", errors="replace")
                reader = csv_mod.reader(StringIO(text_data))
                rows = [", ".join(r) for r in reader if any(r)]
                combined = "\n".join(rows)
                if combined.strip():
                    pages.append((combined, 1))
            except Exception as exc:
                _logger.error("CSV extraction failed for %s: %s", filename, exc)
        else:
            try:
                from openpyxl import load_workbook

                wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
                for sheet_idx, sheet in enumerate(wb.worksheets, start=1):
                    row_texts = []
                    for row in sheet.iter_rows(values_only=True):
                        cells = [str(c) for c in row if c is not None and str(c).strip()]
                        if cells:
                            row_texts.append(", ".join(cells))
                    if row_texts:
                        pages.append(("\n".join(row_texts), sheet_idx))
                wb.close()
            except Exception as exc:
                _logger.error("Excel extraction failed for %s: %s", filename, exc)

        return pages

    def _extract_text(self, data: bytes, filename: str) -> list[tuple[str, int]]:
        """Extract plain text content."""
        try:
            text = data.decode("utf-8", errors="replace").strip()
            if text:
                return [(text, 1)]
        except Exception as exc:
            _logger.error("Text extraction failed for %s: %s", filename, exc)
        return []

    # ------------------------------------------------------------------
    # Chunking
    # ------------------------------------------------------------------

    def _split_into_chunks(
        self,
        pages: list[tuple[str, int]],
        filename: str,
    ) -> List[DocumentChunk]:
        """Split page texts into overlapping chunks using LangChain splitter."""
        try:
            from langchain_text_splitters import RecursiveCharacterTextSplitter

            splitter = RecursiveCharacterTextSplitter(
                chunk_size=self.chunk_size,
                chunk_overlap=self.chunk_overlap,
                length_function=len,
                separators=["\n\n", "\n", ". ", " ", ""],
            )
        except ImportError:
            _logger.warning("langchain-text-splitters not installed; using naive chunking")
            splitter = None

        chunks: List[DocumentChunk] = []
        global_chunk_idx = 0

        for text, page_num in pages:
            if splitter is not None:
                split_texts = splitter.split_text(text)
            else:
                # Naive fallback: fixed-size windows
                split_texts = [
                    text[i: i + self.chunk_size]
                    for i in range(0, len(text), self.chunk_size - self.chunk_overlap)
                    if text[i: i + self.chunk_size].strip()
                ]

            for chunk_text in split_texts:
                if not chunk_text.strip():
                    continue
                chunks.append(
                    DocumentChunk(
                        text=chunk_text.strip(),
                        filename=filename,
                        chunk_index=global_chunk_idx,
                        page_number=page_num,
                        metadata={
                            "source": filename,
                            "page": page_num,
                            "chunk": global_chunk_idx,
                            "source_type": "company_document",
                        },
                    )
                )
                global_chunk_idx += 1

        _logger.info(
            "Document processed: filename=%s pages=%d chunks=%d",
            filename, len(pages), len(chunks),
        )
        return chunks


def get_document_processor(
    chunk_size: Optional[int] = None,
    chunk_overlap: Optional[int] = None,
) -> DocumentProcessor:
    """Return a configured DocumentProcessor instance."""
    return DocumentProcessor(chunk_size=chunk_size, chunk_overlap=chunk_overlap)
